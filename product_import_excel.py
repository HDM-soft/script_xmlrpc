#!/usr/bin/python3

from xmlrpc import client
import openpyxl

# Configuración de conexión
url = 'http://localhost:8269'  # URL del servidor Odoo
dbname = 'db_name'  # Nombre de la base de datos
user = 'admin'  # Usuario de Odoo
pwd = 'admin'  # Contraseña del usuario de Odoo

# Autenticación con Odoo
common = client.ServerProxy(f'{url}/xmlrpc/2/common')  # Conexión al endpoint de autenticación de Odoo
uid = common.authenticate(dbname, user, pwd, {})  # Autenticación con Odoo, obteniendo el ID de usuario
models = client.ServerProxy(f'{url}/xmlrpc/2/object')  # Conexión al endpoint de modelos de Odoo

# Función para crear un atributo y su valor
def crear_atributo_valores(attr_name, valores):
    attr_id = None
    value_ids = []

    # Crear o buscar el atributo
    attr_id = models.execute_kw(dbname, uid, pwd, 'product.attribute', 'search', [[['name', '=', attr_name]]])
    if not attr_id:  # Si el atributo no existe, se crea uno nuevo
        attr_id = models.execute_kw(dbname, uid, pwd, 'product.attribute', 'create', [{'name': attr_name}])
        print(f"[INFO] Atributo creado: {attr_name} con ID {attr_id}")
    else:  # Si existe, se obtiene el ID del atributo existente
        attr_id = attr_id[0]
        print(f"[INFO] Atributo encontrado: {attr_name} con ID {attr_id}")

    # Procesar cada valor y agregarlo al atributo
    for attr_value in valores:
        # Crear o buscar cada valor del atributo
        value_id = models.execute_kw(dbname, uid, pwd, 'product.attribute.value', 'search', [[['attribute_id', '=', attr_id], ['name', '=', attr_value]]])
        if not value_id:  # Si el valor no existe, se crea
            value_id = models.execute_kw(dbname, uid, pwd, 'product.attribute.value', 'create', [{'name': attr_value, 'attribute_id': attr_id}])
            print(f"[INFO] Valor de atributo creado: {attr_value} para el atributo {attr_name} con ID {value_id}")
        else:  # Si existe, se obtiene el ID del valor existente
            value_id = value_id[0]
            print(f"[INFO] Valor de atributo encontrado: {attr_value} para el atributo {attr_name} con ID {value_id}")
        value_ids.append(value_id)  # Se añade el ID del valor a la lista de IDs de valores

    return attr_id, value_ids  # Retorna el ID del atributo y los IDs de sus valores

# Leer atributos desde Excel y crearlos en Odoo
workbook_atr = openpyxl.load_workbook("atributos2.xlsx")  # Cargar archivo Excel con atributos
sheet_atr = workbook_atr.active  # Obtener la hoja activa
for row in sheet_atr.iter_rows(min_row=2):  # Empieza en la segunda fila para omitir encabezados
    attr_name, attr_value = row[0].value, row[1].value  # Leer nombre del atributo y su valor
    crear_atributo_valores(attr_name, [attr_value]) if attr_value else None  # Crear atributo si el valor no es None
workbook_atr.close()  # Cerrar archivo Excel

# Leer productos desde Excel y crear variantes en Odoo
workbook_prod = openpyxl.load_workbook("productos_modificados3.xlsx") 
sheet_prod = workbook_prod.active  # Obtener la hoja activa

for row in sheet_prod.iter_rows(min_row=2):  # Empieza en la segunda fila para omitir encabezados
    nombre_producto = row[0].value  # Nombre del producto
    codigo_barras = row[1].value
    referencia = row[2].value  # Referencia del producto
    tipo_producto = row[3].value
    categoria_producto = row[4].value
    categoria_pdv = row[5].value
    precio_venta = row[6].value
    attr_name = row[8].value  # Nombre del atributo
    attr_values = row[9].value  # Valores del atributo (posiblemente múltiples valores separados por comas)

    # Verificación para evitar valores None
    if not nombre_producto or not referencia or not attr_name or not attr_values:
        print(f"[INFO] Fila incompleta en productos: {nombre_producto}, {referencia}, {attr_name}, {attr_values}. Saltando...")
        continue

    # Separar los valores por coma y quitar espacios en blanco
    valores = [v.strip() for v in attr_values.split(",")]

    # Crear producto principal en product.template si no existe
    template_id = models.execute_kw(dbname, uid, pwd, 'product.template', 'search', [[['name', '=', nombre_producto]]])
    if not template_id:  # Si el producto no existe, se crea uno nuevo
        template_id = models.execute_kw(dbname, uid, pwd, 'product.template', 'create', [{'name': nombre_producto, 'default_code': referencia}])
        print(f"[INFO] Producto creado: {nombre_producto} con referencia {referencia} y ID {template_id}")
    else:  # Si existe, se obtiene el ID del producto existente
        template_id = template_id[0]
        print(f"[INFO] Producto encontrado: {nombre_producto} con referencia {referencia} y ID {template_id}")

    # Crear atributos y valores (múltiples si los hay)
    attr_id, value_ids = crear_atributo_valores(attr_name, valores)

    # Si `attr_id` o `value_ids` está vacío, omitir la creación de la variante
    if attr_id is None or not value_ids:
        continue

    # Agregar el atributo y sus valores en `attribute_line_ids` de product.template
    existing_line = models.execute_kw(dbname, uid, pwd, 'product.template.attribute.line', 'search', [[['product_tmpl_id', '=', template_id], ['attribute_id', '=', attr_id]]])
    if existing_line:  # Si la línea de atributo ya existe, agregar los valores nuevos
        models.execute_kw(dbname, uid, pwd, 'product.template.attribute.line', 'write', [existing_line, {'value_ids': [(4, value_id) for value_id in value_ids]}])
        print(f"[INFO] Valores de atributo {valores} agregados a la línea existente del atributo {attr_name} en el producto {nombre_producto}")
    else:  # Crear una nueva línea de atributo con los valores
        models.execute_kw(dbname, uid, pwd, 'product.template.attribute.line', 'create', [{
            'product_tmpl_id': template_id,
            'attribute_id': attr_id,
            'value_ids': [(4, value_id) for value_id in value_ids]
        }])
        print(f"[INFO] Línea de atributo creada para {attr_name} con los valores {valores} en el producto {nombre_producto}")

    # Asignar referencia (default_code) a cada variante
    variant_ids = models.execute_kw(dbname, uid, pwd, 'product.product', 'search', [[['product_tmpl_id', '=', template_id]]])
    for variant_id in variant_ids:  # Para cada variante creada, asignar el código de referencia
        models.execute_kw(dbname, uid, pwd, 'product.product', 'write', [variant_id, {'default_code': referencia}])
        print(f"[INFO] Referencia {referencia} asignada a la variante con ID {variant_id} del producto {nombre_producto}")

workbook_prod.close()  # Cerrar archivo Excel de productos
print("[INFO] Importación completada.")  # Mensaje final indicando que el proceso ha finalizado
